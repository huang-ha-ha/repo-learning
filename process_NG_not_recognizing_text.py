from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
from datetime import datetime
import shutil
import re
import glob
import math
import sys
import subprocess
from PIL import Image as PILImage

from extract_zip_files import start_extract_zip


# 检查并安装必要的依赖
def install_dependencies():
    try:
        import PILImage
    except ImportError:
        print("正在安装必要的依赖库 Pillow...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pillow"])
            print("Pillow 安装成功！")
            from PIL import Image as PILImage
        except Exception as e:
            print(f"安装 Pillow 失败: {str(e)}")
            print("请手动安装: pip install pillow")
            sys.exit(1)


# 安装依赖
install_dependencies()

# 忽略openpyxl的样式警告
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

# 图片处理设置
IMAGE_HEIGHT = 120  # 图片高度（像素）
IMAGE_MARGIN = 15  # 图片间距（像素）
ROW_HEIGHT = IMAGE_HEIGHT + IMAGE_MARGIN * 2  # 总行高（包含上下边距）


def find_ng_images(sn, data_dir='data'):
    """
    在data目录中查找包含指定SN和"NG"的图片
    """
    matched_images = []

    # 安全处理SN值（可能为数字或字符串）
    sn_str = str(sn).strip() if sn is not None else ""

    if not sn_str:
        return matched_images

    # 构建搜索模式（不区分大小写）
    patterns = [
        f"*{re.escape(sn_str)}*NG*.jpg",
        f"*{re.escape(sn_str)}*NG*.jpeg",
        f"*{re.escape(sn_str)}*NG*.png"
    ]

    # 递归搜索所有匹配的图片
    for pattern in patterns:
        matched_images.extend(glob.glob(os.path.join(data_dir, "**", pattern), recursive=True))

    # 添加不区分大小写的匹配
    if not matched_images:
        for root, dirs, files in os.walk(data_dir):
            for file in files:
                if file.lower().endswith(
                        ('.jpg', '.jpeg', '.png')) and "ng" in file.lower() and sn_str.lower() in file.lower():
                    matched_images.append(os.path.join(root, file))

    return list(set(matched_images))  # 去重


def calculate_image_size(img_path, target_height):
    """
    计算调整后的图片大小（保持宽高比）
    """
    try:
        with PILImage.open(img_path) as img:
            # 保持宽高比调整大小
            width_ratio = target_height / img.height
            new_width = int(img.width * width_ratio)
            return new_width, target_height
    except Exception as e:
        print(f"计算图片大小失败: {str(e)}")
        return 100, 100  # 默认大小


def insert_images_horizontally(worksheet, row_idx, start_col, image_paths):
    """
    将图片水平插入到工作表中（不换行，不显示文件名）
    """
    if not image_paths:
        return {}

    # 设置行高
    worksheet.row_dimensions[row_idx].height = ROW_HEIGHT

    # 存储每列的宽度
    col_widths = {}
    x_offset = 0  # 水平偏移量

    for idx, img_path in enumerate(image_paths):
        try:
            # 计算图片大小
            img_width, img_height = calculate_image_size(img_path, IMAGE_HEIGHT)

            # 确定当前列
            current_col = start_col + idx

            # 记录列宽
            col_widths[current_col] = img_width

            # 加载图片
            img = Image(img_path)
            img.width = img_width
            img.height = img_height

            # 设置图片位置
            img.anchor = f"{get_column_letter(current_col)}{row_idx}"
            img.left = x_offset
            worksheet.add_image(img)

            # 更新水平偏移量
            x_offset += img_width + IMAGE_MARGIN

        except Exception as e:
            print(f"插入图片 {img_path} 时出错: {str(e)}")

    return col_widths


def apply_image_dimensions(worksheet, col_widths):
    """
    应用图片尺寸到工作表列宽
    """
    # 像素到Excel列宽单位的转换因子
    PIXELS_TO_EXCEL_UNITS = 0.14

    for col_idx, width_px in col_widths.items():
        # 计算列宽（Excel单位）
        col_width = max(15, width_px * PIXELS_TO_EXCEL_UNITS)
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = col_width


def format_excel(worksheet):
    """
    格式化Excel表格，使其更易读
    """
    # 设置基础列宽
    column_widths = {
        'A': 25,  # SN
        'B': 25,  # QPL-Station Name
        'C': 15,  # Gantry
        'D': 20,  # Time(end)
        'E': 25,  # locate picture
    }

    # 应用基础列宽
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    # 设置标题样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用标题样式
    max_col = worksheet.max_column
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据行样式
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

            # 设置日期格式
            if col == 4 and isinstance(cell.value, datetime):
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # 冻结首行
    worksheet.freeze_panes = 'A2'


def extract_columns(argv):
    # 创建result目录（如果不存在）
    result_dir = "result"
    os.makedirs(result_dir, exist_ok=True)

    # 创建图片目录
    images_dir = os.path.join(result_dir, "images")
    os.makedirs(images_dir, exist_ok=True)

    # 输入文件路径
    input_file = argv[1]

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件不存在: {input_file}")

    # 加载工作簿
    wb = load_workbook(input_file, data_only=True)

    # 查找包含"不良明细"的工作表
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "不良明细" in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        available_sheets = "\n".join(wb.sheetnames)
        raise ValueError(f"未找到包含'不良明细'的工作表。可用工作表有：\n{available_sheets}")

    print(f"找到目标工作表: {target_sheet}")

    # 获取目标工作表
    sheet = wb[target_sheet]

    # 查找列索引
    header_row = 1  # 假设标题在第一行
    col_index = {}
    required_columns = ['SN', 'Station Name', 'Time End']

    for cell in sheet[header_row]:
        if cell.value in required_columns:
            col_index[cell.value] = cell.column_letter

    # 检查是否找到所有列
    missing_cols = [col for col in required_columns if col not in col_index]

    if missing_cols:
        # 尝试大小写不敏感匹配
        col_index = {}
        for cell in sheet[header_row]:
            cell_value = str(cell.value).lower() if cell.value else ""
            for req_col in required_columns:
                if req_col.lower() == cell_value:
                    col_index[req_col] = cell.column_letter

        # 再次检查
        missing_cols = [col for col in required_columns if col not in col_index]

        if missing_cols:
            available_cols = [cell.value for cell in sheet[header_row] if cell.value]
            raise ValueError(f"以下列在目标工作表中不存在: {', '.join(missing_cols)}\n"
                             f"可用列: {', '.join(filter(None, available_cols))}")

    # 创建新工作簿
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "不良明细汇总"

    # 添加新标题（按指定顺序）
    new_headers = [
        'SN',
        'QPL-Station Name',
        'Gantry',
        'Time(end)',
        'locate picture',
        'NG picture'
    ]

    # 写入新标题
    for col_idx, header in enumerate(new_headers, start=1):
        new_sheet.cell(row=1, column=col_idx, value=header)

    # 存储所有列的最大宽度
    all_col_widths = {}

    # 复制数据
    row_idx = 2  # 数据从第2行开始
    for src_row in range(header_row + 1, sheet.max_row + 1):
        # 获取所需列的值
        sn_value = sheet[f"{col_index['SN']}{src_row}"].value
        station_value = sheet[f"{col_index['Station Name']}{src_row}"].value
        time_value = sheet[f"{col_index['Time End']}{src_row}"].value

        # 跳过空行
        if not any([sn_value, station_value, time_value]):
            continue

        try:
            # 写入基础数据
            new_sheet.cell(row=row_idx, column=1, value=sn_value)
            new_sheet.cell(row=row_idx, column=2, value=station_value)
            new_sheet.cell(row=row_idx, column=3, value=None)  # Gantry
            new_sheet.cell(row=row_idx, column=4, value=time_value)
            new_sheet.cell(row=row_idx, column=5, value=None)  # locate picture

            # 查找NG图片
            ng_images = []
            if sn_value:
                ng_images = find_ng_images(sn_value)

                if ng_images:
                    print(f"为SN {sn_value} 找到 {len(ng_images)} 张NG图片")

                    # 复制图片到结果目录
                    copied_images = []
                    for img_path in ng_images:
                        try:
                            img_name = os.path.basename(img_path)
                            dest_path = os.path.join(images_dir, img_name)
                            shutil.copy2(img_path, dest_path)
                            copied_images.append(dest_path)
                        except Exception as e:
                            print(f"复制图片失败: {str(e)}")

                    # 插入图片到工作表（水平排列）
                    if copied_images:
                        # 在NG picture列开始插入图片
                        col_widths = insert_images_horizontally(new_sheet, row_idx, 6, copied_images)

                        # 更新全局列宽记录
                        for col_idx, width in col_widths.items():
                            if col_idx not in all_col_widths or width > all_col_widths[col_idx]:
                                all_col_widths[col_idx] = width

            # 移动到下一行
            row_idx += 1
        except Exception as e:
            print(f"警告: 处理行 {src_row} 时出错 - {str(e)}")
            row_idx += 1

    # 应用图片尺寸到列宽
    if all_col_widths:
        apply_image_dimensions(new_sheet, all_col_widths)

    # 格式化Excel表格
    if new_sheet.max_row > 1:  # 确保有数据行
        format_excel(new_sheet)

    # 保存新工作簿到result目录
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(result_dir, f"不良明细汇总_{timestamp}.xlsx")
    new_wb.save(output_file)

    print(f"成功创建新文件: {output_file}")
    print(f"处理了 {row_idx - 2} 条记录")
    print(f"源工作表: {target_sheet}")


if __name__ == "__main__":
    try:
        start_extract_zip()
        # 确保data目录存在
        if not os.path.exists("data"):
            print("警告: data目录不存在，将跳过图片搜索")

        extract_columns(sys.argv)
    except Exception as e:
        print(f"发生错误: {str(e)}")
        import traceback

        traceback.print_exc()
        print("请确保:")
        print("1. 原始Excel文件存在且路径正确")
        print("2. 文件没有被其他程序占用")
        print("3. 包含'不良明细'的工作表存在")
        print("4. 工作表中包含SN, Station Name, Time End三列")

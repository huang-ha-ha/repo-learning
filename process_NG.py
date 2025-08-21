from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
from datetime import datetime
import shutil
import re
import glob
import sys
import subprocess
from PIL import Image as PILImage
import traceback
import argparse

from extract_zip_files import start_extract_zip


# 检查并安装必要的依赖
def install_dependencies():
    dependencies = ['pillow']
    installed = False

    # 检查并安装Pillow
    try:
        import PIL
    except ImportError:
        print("正在安装必要的依赖库 Pillow...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pillow"])
            print("Pillow 安装成功！")
            installed = True
        except Exception as e:
            print(f"安装 Pillow 失败: {str(e)}")
            print("请手动安装: pip install pillow")
            sys.exit(1)

    if installed:
        print("所有依赖已成功安装，请重新运行脚本")
        return True
    return True


# 安装依赖
if not install_dependencies():
    print("依赖安装失败，请手动安装必要组件")
    sys.exit(1)

# 忽略openpyxl的样式警告
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

# 图片处理设置
IMAGE_HEIGHT = 120  # 图片高度（像素）
IMAGE_MARGIN = 15  # 图片间距（像素）

# 设备类型
DEVICE_TYPES = ['1100', '660', '1174', '639']


def find_sn_folders(sn, data_dir='data'):
    """
    在data目录中查找包含指定SN的文件夹
    """
    sn_str = str(sn).strip() if sn is not None else ""
    if not sn_str:
        return []

    # 检查data_dir是否存在
    if not os.path.exists(data_dir):
        print(f"目录 {data_dir} 不存在，跳过文件夹搜索")
        return []

    # 查找包含SN的文件夹
    matched_folders = []
    for root, dirs, files in os.walk(data_dir):
        for dir_name in dirs:
            if sn_str in dir_name:
                matched_folders.append(os.path.join(root, dir_name))

    return matched_folders


def find_all_images_in_folder(folder_path):
    """
    在指定文件夹中查找所有图片
    """
    all_images = []
    min_file_size = 10 * 1024  # 10KB最小文件大小

    if not os.path.exists(folder_path):
        return all_images

    # 查找所有图片文件
    for ext in ['*.jpg', '*.jpeg', '*.png']:
        for img_path in glob.glob(os.path.join(folder_path, ext)):
            # 检查文件大小
            if os.path.getsize(img_path) < min_file_size:
                print(f"跳过小文件: {os.path.basename(img_path)} (大小: {os.path.getsize(img_path) / 1024:.1f}KB)")
                continue

            # 检查是否为有效图片
            try:
                with PILImage.open(img_path) as img:
                    img.verify()  # 验证图片完整性
            except Exception as e:
                print(f"跳过损坏图片: {os.path.basename(img_path)} - {str(e)}")
                continue

            all_images.append(img_path)

    return all_images


def filter_ng_images(images, device_type):
    """
    根据设备类型过滤NG图片
    """
    ng_images = []
    src_images = []

    for img_path in images:
        img_name = os.path.basename(img_path).lower()

        # 检查是否包含NG
        if 'ng' in img_name:
            # 检查是否包含src
            if 'src' in img_name:
                src_images.append(img_path)
            else:
                ng_images.append(img_path)

    # 根据设备类型处理
    if device_type in ['1100', '660']:
        # 删除包含src的NG图片
        ng_images = [img for img in ng_images if img not in src_images]

        # 按修改时间排序，取最后几张
        ng_images.sort(key=lambda x: os.path.getmtime(x))

    elif device_type == '1174':
        # 删除包含src的NG图片
        ng_images = [img for img in ng_images if img not in src_images]

        # 按修改时间排序，取最后几张
        ng_images.sort(key=lambda x: os.path.getmtime(x))

    elif device_type == '639':
        # 删除包含src的NG图片
        ng_images = [img for img in ng_images if img not in src_images]

        # 按修改时间排序，取最后几张
        ng_images.sort(key=lambda x: os.path.getmtime(x))

    return ng_images, src_images


def filter_ok_images(images):
    """
    过滤OK图片
    """
    ok_images = []

    for img_path in images:
        img_name = os.path.basename(img_path).lower()

        # 检查是否包含OK
        if 'ok' in img_name:
            ok_images.append(img_path)

    return ok_images


def process_1100_660(ng_images, ok_images, folder_path):
    """
    处理1100和660类型的图片
    """
    remarks = []
    locate_image = None

    if len(ng_images) == 0:
        remarks.append("Error: 未找到NG图片")
        return ng_images, locate_image, "\n".join(remarks)

    # 检查是否全是src图片
    if len(ng_images) == 0 and len(ok_images) > 0:
        remarks.append("Error: 只有OK图片，没有NG图片")
        return ng_images, locate_image, "\n".join(remarks)

    # 如果只有一张NG图片
    if len(ng_images) == 1:
        ng_image = ng_images[0]
        ng_name = os.path.basename(ng_image)

        # 提取NG前的名称部分
        match = re.search(r'(\d{14}-Station\d+)', ng_name)
        if match:
            ng_prefix = match.group(1)

            # 查找对应的OK图片
            for ok_image in ok_images:
                ok_name = os.path.basename(ok_image)
                if ng_prefix in ok_name and 'ok' in ok_name.lower():
                    locate_image = ok_image
                    break

            # 如果没有找到完全匹配的OK图片，查找类似的
            if not locate_image:
                for ok_image in ok_images:
                    ok_name = os.path.basename(ok_image)
                    # 检查是否有类似的OK图片（例如不同的Station编号）
                    if re.search(r'\d{14}-Station\d+-OK', ok_name, re.IGNORECASE):
                        remarks.append(f"Failed: 未找到完全匹配的OK图片，但有类似的OK图片: {os.path.basename(ok_image)}")
                        break

        return [ng_image], locate_image, "\n".join(remarks) if remarks else ""

    # 如果有多张NG图片，取最后2张
    if len(ng_images) > 1:
        selected_ng_images = ng_images[-2:]

        # 检查两张NG图片的名称是否一致
        ng_name1 = os.path.basename(selected_ng_images[0])
        ng_name2 = os.path.basename(selected_ng_images[1])

        match1 = re.search(r'(\d{14}-Station\d+)', ng_name1)
        match2 = re.search(r'(\d{14}-Station\d+)', ng_name2)

        if match1 and match2:
            ng_prefix1 = match1.group(1)
            ng_prefix2 = match2.group(1)

            if ng_prefix1 != ng_prefix2:
                remarks.append(f"Failed: 两张NG图片名称不一致: {ng_prefix1} vs {ng_prefix2}")
            else:
                # 查找对应的OK图片
                for ok_image in ok_images:
                    ok_name = os.path.basename(ok_image)
                    if ng_prefix1 in ok_name and 'ok' in ok_name.lower():
                        locate_image = ok_image
                        break

                # 如果没有找到完全匹配的OK图片，查找类似的
                if not locate_image:
                    for ok_image in ok_images:
                        ok_name = os.path.basename(ok_image)
                        # 检查是否有类似的OK图片（例如不同的Station编号）
                        if re.search(r'\d{14}-Station\d+-OK', ok_name, re.IGNORECASE):
                            remarks.append(
                                f"Failed: 未找到完全匹配的OK图片，但有类似的OK图片: {os.path.basename(ok_image)}")
                            break

        return selected_ng_images, locate_image, "\n".join(remarks) if remarks else ""

    return ng_images, locate_image, "\n".join(remarks) if remarks else ""


def process_1174(ng_images, ok_images, folder_path):
    """
    处理1174类型的图片
    """
    remarks = []
    locate_image = None

    if len(ng_images) == 0:
        remarks.append("Error: 未找到NG图片")
        return ng_images, locate_image, "\n".join(remarks)

    # 如果只有一张NG图片
    if len(ng_images) == 1:
        ng_image = ng_images[0]
        ng_name = os.path.basename(ng_image)

        # 提取NG前的名称部分
        match = re.search(r'(Pose\d+_\d{12})', ng_name)
        if match:
            ng_prefix = match.group(1)

            # 查找对应的OK图片
            for ok_image in ok_images:
                ok_name = os.path.basename(ok_image)
                if ng_prefix in ok_name and 'ok' in ok_name.lower():
                    locate_image = ok_image
                    break

            # 如果没有找到完全匹配的OK图片，查找类似的
            if not locate_image:
                for ok_image in ok_images:
                    ok_name = os.path.basename(ok_image)
                    # 检查是否有类似的OK图片（例如不同的Pose编号）
                    if re.search(r'Pose\d+_\d{12}-OK', ok_name, re.IGNORECASE):
                        remarks.append(f"Failed: 未找到完全匹配的OK图片，但有类似的OK图片: {os.path.basename(ok_image)}")
                        break

        return [ng_image], locate_image, "\n".join(remarks) if remarks else ""

    # 如果有多张NG图片，取最后2张
    if len(ng_images) > 1:
        selected_ng_images = ng_images[-2:]

        # 检查两张NG图片的名称是否一致
        ng_name1 = os.path.basename(selected_ng_images[0])
        ng_name2 = os.path.basename(selected_ng_images[1])

        match1 = re.search(r'(Pose\d+_\d{12})', ng_name1)
        match2 = re.search(r'(Pose\d+_\d{12})', ng_name2)

        if match1 and match2:
            ng_prefix1 = match1.group(1)
            ng_prefix2 = match2.group(1)

            if ng_prefix1 != ng_prefix2:
                remarks.append(f"Failed: 两张NG图片名称不一致: {ng_prefix1} vs {ng_prefix2}")
            else:
                # 查找对应的OK图片
                for ok_image in ok_images:
                    ok_name = os.path.basename(ok_image)
                    if ng_prefix1 in ok_name and 'ok' in ok_name.lower():
                        locate_image = ok_image
                        break

                # 如果没有找到完全匹配的OK图片，查找类似的
                if not locate_image:
                    for ok_image in ok_images:
                        ok_name = os.path.basename(ok_image)
                        # 检查是否有类似的OK图片（例如不同的Pose编号）
                        if re.search(r'Pose\d+_\d{12}-OK', ok_name, re.IGNORECASE):
                            remarks.append(
                                f"Failed: 未找到完全匹配的OK图片，但有类似的OK图片: {os.path.basename(ok_image)}")
                            break

        return selected_ng_images, locate_image, "\n".join(remarks) if remarks else ""

    return ng_images, locate_image, "\n".join(remarks) if remarks else ""


def process_639(ng_images, ok_images, folder_path):
    """
    处理639类型的图片
    """
    remarks = []
    locate_image = None

    if len(ng_images) == 0:
        remarks.append("Error: 未找到NG图片")
        return ng_images, locate_image, "\n".join(remarks)

    # 如果有多张NG图片，取最后2张
    if len(ng_images) > 1:
        ng_images = ng_images[-2:]

    # 如果有OK图片，选择第一张作为locate image
    if ok_images:
        locate_image = ok_images[0]

    return ng_images, locate_image, "\n".join(remarks) if remarks else ""


def process_images_by_device_type(device_type, folder_path):
    """
    根据设备类型处理图片
    """
    # 查找所有图片
    all_images = find_all_images_in_folder(folder_path)

    # 检查文件夹是否为空
    if not all_images:
        return [], None, "Error: 文件夹为空"

    # 过滤NG图片和OK图片
    ng_images, src_images = filter_ng_images(all_images, device_type)
    ok_images = filter_ok_images(all_images)

    # 检查是否全是src图片
    if len(ng_images) == 0 and len(src_images) > 0:
        return [], None, "Error: 只有包含src的NG图片"

    # 检查是否没有NG图片但有OK图片
    if len(ng_images) == 0 and len(ok_images) > 0:
        return [], None, "Error: 只有OK图片，没有NG图片"

    # 检查是否没有NG图片也没有OK图片
    if len(ng_images) == 0 and len(ok_images) == 0:
        return [], None, "Error: 未找到包含NG或OK的图片"

    # 根据设备类型调用不同的处理函数
    if device_type in ['1100', '660']:
        return process_1100_660(ng_images, ok_images, folder_path)
    elif device_type == '1174':
        return process_1174(ng_images, ok_images, folder_path)
    elif device_type == '639':
        return process_639(ng_images, ok_images, folder_path)

    return [], None, "Error: 未知的设备类型"


def calculate_image_size(img_path, target_height):
    """
    计算调整后的图片大小（保持宽高比）
    """
    try:
        with PILImage.open(img_path) as img:
            # 保持宽高比调整大小
            width_ratio = target_height / img.height
            new_width = int(img.width * width_ratio)
            return new_width, target_height
    except Exception as e:
        print(f"计算图片大小失败: {str(e)}")
        return 100, 100  # 默认大小


def insert_images_horizontally(worksheet, row_idx, start_col, image_paths):
    """
    精确控制图片行高度，消除多余空白
    """
    if not image_paths:
        return {}, []

    col_widths = {}
    max_img_height = 0
    images_added = []

    # 先计算所有图片的尺寸
    img_sizes = []
    for img_path in image_paths:
        try:
            width, height = calculate_image_size(img_path, IMAGE_HEIGHT)
            img_sizes.append((width, height))
            if height > max_img_height:
                max_img_height = height
        except Exception as e:
            print(f"计算图片尺寸失败: {str(e)}")
            continue

    # 设置行高（关键修改：使用精确计算方式）
    if max_img_height > 0:
        # Excel行高单位换算：1单位 ≈ 1/7.2毫米 ≈ 1.33像素
        # 精确计算公式：行高 = (图片高度 + 上边距) / 1.33
        exact_row_height = (max_img_height + IMAGE_MARGIN) / 1.33
        worksheet.row_dimensions[row_idx].height = exact_row_height
        print(f"精确设置行 {row_idx} 高度: {exact_row_height:.2f} (像素高度: {max_img_height}+{IMAGE_MARGIN})")

    # 插入图片
    for idx, (img_path, (width, height)) in enumerate(zip(image_paths, img_sizes)):
        try:
            current_col = start_col + idx
            col_widths[current_col] = width

            img = Image(img_path)
            img.width = width
            img.height = height

            # 关键修改：使用单元格锚定方式（消除额外空白）
            cell_anchor = f"{get_column_letter(current_col)}{row_idx}"
            img.anchor = cell_anchor
            worksheet.add_image(img)
            images_added.append(img)
        except Exception as e:
            print(f"插入图片失败: {str(e)}")

    return col_widths, images_added


def apply_image_dimensions(worksheet, col_widths):
    """
    应用图片尺寸到工作表列宽
    """
    # 像素到Excel列宽单位的转换因子
    PIXELS_TO_EXCEL_UNITS = 0.14

    for col_idx, width_px in col_widths.items():
        # 计算列宽（Excel单位）
        col_width = max(15, width_px * PIXELS_TO_EXCEL_UNITS)
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = col_width
        print(f"设置列 {col_letter} 宽度为: {col_width} (基于图片宽度: {width_px} 像素)")


def format_excel(worksheet):
    """
    格式化Excel表格，使其更易读，并优化行高
    """
    # 设置基础列宽
    column_widths = {
        'A': 25,  # SN
        'B': 25,  # QPL-Station Name
        'C': 15,  # Gantry
        'D': 20,  # Time(end)
        'E': 25,  # Locate picture
        'F': 25,  # NG picture
        'G': 25,  # NG picture1
        'H': 40,  # Remark
    }

    # 应用基础列宽
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    # 设置标题样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用标题样式
    max_col = worksheet.max_column
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据行样式
    min_row_height = 20  # 默认最小行高（像素）

    # 设置行高 - 只对没有图片的行设置默认行高
    # 有图片的行已经在插入时设置了精确高度
    for row in range(2, worksheet.max_row + 1):
        # 检查该行是否有图片
        has_images = False
        for image in worksheet._images:
            # 从图片的anchor字符串中提取行号
            anchor_str = image.anchor
            if isinstance(anchor_str, str):
                match = re.search(r'\d+$', anchor_str)
                if match:
                    image_row = int(match.group())
                    if image_row == row:
                        has_images = True
                        break

        # 如果没有图片，设置默认行高
        if not has_images:
            worksheet.row_dimensions[row].height = min_row_height

        # 设置单元格样式
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)

            # 设置日期格式
            if col == 4 and isinstance(cell.value, datetime):
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # 冻结首行
    worksheet.freeze_panes = 'A2'


def extract_columns(device_type, input_file):
    # 创建result目录（如果不存在）
    result_dir = "result"
    os.makedirs(result_dir, exist_ok=True)

    # 创建图片目录
    images_dir = os.path.join(result_dir, "images")
    os.makedirs(images_dir, exist_ok=True)

    print(f"使用文件: {input_file}")

    # 检查文件是否存在
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"输入文件不存在: {input_file}")

    # 加载工作簿
    wb = load_workbook(input_file, data_only=True)

    # 查找包含"不良明细"的工作表
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "不良明细" in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        available_sheets = "\n".join(wb.sheetnames)
        raise ValueError(f"未找到包含'不良明细'的工作表。可用工作表有：\n{available_sheets}")

    print(f"找到目标工作表: {target_sheet}")

    # 获取目标工作表
    sheet = wb[target_sheet]

    # 查找列索引
    header_row = 1  # 假设标题在第一行
    col_index = {}
    required_columns = ['SN', 'Station Name', 'Time End']

    for cell in sheet[header_row]:
        if cell.value in required_columns:
            col_index[cell.value] = cell.column_letter

    # 检查是否找到所有列
    missing_cols = [col for col in required_columns if col not in col_index]

    if missing_cols:
        # 尝试大小写不敏感匹配
        col_index = {}
        for cell in sheet[header_row]:
            cell_value = str(cell.value).lower() if cell.value else ""
            for req_col in required_columns:
                if req_col.lower() == cell_value:
                    col_index[req_col] = cell.column_letter

        # 再次检查
        missing_cols = [col for col in required_columns if col not in col_index]

        if missing_cols:
            available_cols = [cell.value for cell in sheet[header_row] if cell.value]
            raise ValueError(f"以下列在目标工作表中不存在: {', '.join(missing_cols)}\n"
                             f"可用列: {', '.join(filter(None, available_cols))}")

    # 创建新工作簿
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "不良明细汇总"

    # 添加新标题（按指定顺序）
    new_headers = [
        'SN',
        'QPL-Station Name',
        'Gantry',
        'Time(end)',
        'Locate picture',
        'NG picture',
        'NG picture1',
        'Remark'
    ]

    # 写入新标题
    for col_idx, header in enumerate(new_headers, start=1):
        new_sheet.cell(row=1, column=col_idx, value=header)

    # 存储所有列的最大宽度
    all_col_widths = {}

    # 复制数据
    row_idx = 2  # 数据从第2行开始
    for src_row in range(header_row + 1, sheet.max_row + 1):
        # 获取所需列的值
        sn_value = sheet[f"{col_index['SN']}{src_row}"].value
        station_value = sheet[f"{col_index['Station Name']}{src_row}"].value
        time_value = sheet[f"{col_index['Time End']}{src_row}"].value

        # 跳过空行
        if not any([sn_value, station_value, time_value]):
            continue

        try:
            # 写入基础数据
            new_sheet.cell(row=row_idx, column=1, value=sn_value)
            new_sheet.cell(row=row_idx, column=2, value=station_value)
            new_sheet.cell(row=row_idx, column=3, value=None)  # Gantry
            new_sheet.cell(row=row_idx, column=4, value=time_value)
            new_sheet.cell(row=row_idx, column=5, value=None)  # Locate picture

            # 初始化Remark列
            remark = ""

            # 查找SN对应的文件夹
            sn_folders = find_sn_folders(sn_value)

            if len(sn_folders) > 1:
                # 多个文件夹匹配，记录错误
                folder_names = ", ".join([os.path.basename(f) for f in sn_folders])
                remark = f"Error: 多个文件夹匹配 - {folder_names}"
                new_sheet.cell(row=row_idx, column=8, value=remark)
                print(f"为SN {sn_value} 找到多个匹配文件夹: {folder_names}")
            elif len(sn_folders) == 1:
                # 找到一个文件夹，在其中处理图片
                folder_path = sn_folders[0]
                print(f"为SN {sn_value} 找到匹配文件夹: {os.path.basename(folder_path)}")

                # 根据设备类型处理图片
                ng_images, locate_image, process_remark = process_images_by_device_type(device_type, folder_path)

                # 复制图片到结果目录
                copied_ng_images = []
                for img_path in ng_images:
                    try:
                        img_name = os.path.basename(img_path)
                        dest_path = os.path.join(images_dir, img_name)
                        shutil.copy2(img_path, dest_path)
                        copied_ng_images.append(dest_path)
                    except Exception as e:
                        print(f"复制NG图片失败: {str(e)}")

                copied_locate_image = None
                if locate_image:
                    try:
                        img_name = os.path.basename(locate_image)
                        dest_path = os.path.join(images_dir, img_name)
                        shutil.copy2(locate_image, dest_path)
                        copied_locate_image = dest_path
                    except Exception as e:
                        print(f"复制定位图片失败: {str(e)}")

                # 插入图片到工作表
                if copied_ng_images:
                    # 最多显示两张图片
                    max_images = min(2, len(copied_ng_images))
                    images_to_insert = copied_ng_images[:max_images]

                    # 插入第一张图片到NG picture列
                    if len(images_to_insert) >= 1:
                        col_widths1, _ = insert_images_horizontally(new_sheet, row_idx, 6, [images_to_insert[0]])
                        # 更新全局列宽记录
                        for col_idx, width in col_widths1.items():
                            if col_idx not in all_col_widths or width > all_col_widths[col_idx]:
                                all_col_widths[col_idx] = width

                    # 插入第二张图片到NG picture1列
                    if len(images_to_insert) >= 2:
                        col_widths2, _ = insert_images_horizontally(new_sheet, row_idx, 7, [images_to_insert[1]])
                        # 更新全局列宽记录
                        for col_idx, width in col_widths2.items():
                            if col_idx not in all_col_widths or width > all_col_widths[col_idx]:
                                all_col_widths[col_idx] = width

                # 插入定位图片到Locate picture列
                if copied_locate_image:
                    col_widths_loc, _ = insert_images_horizontally(new_sheet, row_idx, 5, [copied_locate_image])
                    # 更新全局列宽记录
                    for col_idx, width in col_widths_loc.items():
                        if col_idx not in all_col_widths or width > all_col_widths[col_idx]:
                            all_col_widths[col_idx] = width

                # 添加处理备注
                if process_remark:
                    if remark:
                        remark += "\n" + process_remark
                    else:
                        remark = process_remark
                    new_sheet.cell(row=row_idx, column=8, value=remark)
            else:
                # 没有找到匹配的文件夹
                remark = "Error: 未找到包含SN的文件夹"
                new_sheet.cell(row=row_idx, column=8, value=remark)
                print(f"为SN {sn_value} 未找到匹配文件夹")

            # 移动到下一行
            row_idx += 1
        except Exception as e:
            print(f"警告: 处理行 {src_row} 时出错 - {str(e)}")
            traceback.print_exc()
            row_idx += 1

    # 应用图片尺寸到列宽
    if all_col_widths:
        apply_image_dimensions(new_sheet, all_col_widths)

    # 格式化Excel表格
    if new_sheet.max_row > 1:  # 确保有数据行
        format_excel(new_sheet)

    # 保存新工作簿到result目录
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(result_dir, f"不良明细汇总_{device_type}_{timestamp}.xlsx")
    new_wb.save(output_file)

    print(f"成功创建新文件: {output_file}")
    print(f"处理了 {row_idx - 2} 条记录")
    print(f"源工作表: {target_sheet}")


if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='处理不良明细数据')
    parser.add_argument('device_type', choices=DEVICE_TYPES, help='设备类型: 1100, 660, 1174, 639')
    parser.add_argument('input_file', help='输入Excel文件路径')
    args = parser.parse_args()

    try:
        start_extract_zip()
        # 确保data目录存在
        if not os.path.exists("data"):
            print("警告: data目录不存在，将跳过图片搜索")
            os.makedirs("data", exist_ok=True)

        extract_columns(args.device_type, args.input_file)
    except Exception as e:
        print(f"发生错误: {str(e)}")
        traceback.print_exc()
        print("请确保:")
        print("1. 原始Excel文件存在且路径正确")
        print("2. 文件没有被其他程序占用")
        print("3. 包含'不良明细'的工作表存在")
        print("4. 工作表中包含SN, Station Name, Time End三列")
